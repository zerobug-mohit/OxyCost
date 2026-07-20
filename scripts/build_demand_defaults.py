#!/usr/bin/env python3
"""Build src/data/demand.json from the case-mix demand workbook.

Reads (all local; the .xlsx stays git-ignored, only the JSON ships):
  - "Scalar Input"            -> 18 wards' base case profiles (flow/duration/mix),
                                 seasonality factors, and scalars (mins/day, MT conv, surge)
  - "Facility strata"         -> 25 tranches (State x Type x admission band) + O2/admission factor
  - "Facility Extrapolated (809)" -> per-facility monthly demand + tranche  (extrapolated set)
  - "Total Facility Output (Y1)"  -> per-facility annual demand + label     (sampled set)

Emits per-district demand: a fixed "sampled" part (ward case-mix) and a per-tranche
"extrapolated" part, plus a per-facility list (name + annual demand) for the drill-down.
Facility NAMES are included (tool-owner decision); admissions counts are NOT shipped.

Run:  python scripts/build_demand_defaults.py
"""
import json
import os
from collections import defaultdict

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
XLSX = os.path.join(ROOT, "Demand_estimation_v9_15-07_Case mix method.xlsx")
OUT = os.path.join(ROOT, "src", "data", "demand.json")

# 18 ward sections in "Scalar Input": header row -> {flow,duration,mix} at fixed offsets.
WARD_HEADER_ROWS = list(range(21, 21 + 14 * 18, 14))  # 21,35,...,259

# Month index (0=Nov-25 ... 11=Oct-26) -> season key, matching the workbook's flags.
MONTH_SEASON = [
    "autumn", "winter", "winter", "winter",   # Nov, Dec, Jan, Feb
    "summer", "summer", "summer",             # Mar, Apr, May
    "monsoon", "monsoon", "monsoon",          # Jun, Jul, Aug
    "autumn", "autumn",                       # Sep, Oct
]
MONTH_LABELS = ["Nov", "Dec", "Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct"]


def num(v, default=0.0):
    try:
        return float(v)
    except (TypeError, ValueError):
        return default


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    si = wb["Scalar Input"]

    # ---- Wards (base = column I / col 9) ----
    wards = {}
    for h in WARD_HEADER_ROWS:
        label = si.cell(h, 1).value
        prefix_cell = si.cell(h + 2, 5).value  # e.g. "hdu_flow_c1"
        if not prefix_cell:
            continue
        key = str(prefix_cell).split("_flow")[0]
        wards[key] = {
            "label": str(label).strip(),
            "flow": [num(si.cell(h + 2, 9).value), num(si.cell(h + 3, 9).value), num(si.cell(h + 4, 9).value)],
            "duration": [num(si.cell(h + 6, 9).value), num(si.cell(h + 7, 9).value), num(si.cell(h + 8, 9).value)],
            "mix": [num(si.cell(h + 10, 9).value), num(si.cell(h + 11, 9).value), num(si.cell(h + 12, 9).value)],
        }
    assert len(wards) == 18, f"expected 18 wards, got {len(wards)}"

    # ---- Seasonality + scalars ----
    seasonality = {
        "winter": num(si.cell(14, 9).value, 1.25),
        "summer": num(si.cell(15, 9).value, 0.75),
        "monsoon": num(si.cell(16, 9).value, 1.05),
        "autumn": num(si.cell(17, 9).value, 1.25),
    }
    scalars = {
        "minsPerDay": num(si.cell(8, 9).value, 1440),
        "mtConversion": num(si.cell(9, 9).value, 750000),
        "pandemicSurge": num(si.cell(10, 9).value, 5),
    }

    # ---- 25 tranches from "Facility strata" ----
    fs = wb["Facility strata"]
    tranches = []
    for r in range(5, fs.max_row + 1):
        state = fs.cell(r, 1).value
        if not state:
            continue
        tranches.append({
            "state": str(state).strip(),
            "type": str(fs.cell(r, 2).value).strip(),
            "band": str(fs.cell(r, 3).value).strip(),
            "upperBound": num(fs.cell(r, 4).value),
            "label": str(fs.cell(r, 5).value).strip(),
            "factor": num(fs.cell(r, 6).value),
        })

    # ---- District aggregates + per-facility rows ----
    # facilities: (state,district) -> [{name, mt, tr}]  (tr = tranche label, or
    # "_sampled" for the ward-based measured set). Names ARE shipped (per the
    # tool owner's decision) so the UI can drill district -> strata -> facility.
    facilities = defaultdict(list)   # (state,district) -> [ {name, mt, tr} ]

    # Extrapolated: annual = sum of monthly (cols G..R = 7..18), grouped by (state, district, tranche label)
    fe = wb["Facility Extrapolated (809)"]
    extrap = defaultdict(lambda: defaultdict(float))   # (state,district) -> {trancheLabel: annualMT}
    counts = defaultdict(int)
    for r in range(5, fe.max_row + 1):
        state = fe.cell(r, 1).value
        if not state:
            continue
        state = str(state).strip()
        district = str(fe.cell(r, 2).value).strip()
        name = str(fe.cell(r, 4).value or "").strip()
        tr = str(fe.cell(r, 6).value).strip()
        annual = sum(num(fe.cell(r, c).value) for c in range(7, 19))
        extrap[(state, district)][tr] += annual
        counts[(state, district)] += 1
        if annual > 0:
            facilities[(state, district)].append({"name": name, "mt": round(annual, 4), "tr": tr})

    # Sampled: annual = S (col 19) from Total Facility Output where label != 'Extrapolated'
    tfo = wb["Total Facility Output (Y1)"]
    sampled = defaultdict(float)
    for r in range(5, tfo.max_row + 1):
        state = tfo.cell(r, 1).value
        if not state:
            continue
        state = str(state).strip()
        label = str(tfo.cell(r, 5).value).strip()
        district = str(tfo.cell(r, 2).value).strip()
        name = str(tfo.cell(r, 4).value or "").strip()
        if label.lower() != "extrapolated":
            annual = num(tfo.cell(r, 19).value)
            sampled[(state, district)] += annual
            counts[(state, district)] += 1
            if annual > 0:
                facilities[(state, district)].append({"name": name, "mt": round(annual, 4), "tr": "_sampled"})

    districts = defaultdict(dict)
    all_keys = set(extrap.keys()) | set(sampled.keys())
    for (state, district) in sorted(all_keys):
        facs = sorted(facilities.get((state, district), []), key=lambda x: -x["mt"])
        districts[state][district] = {
            "sampledMT": round(sampled.get((state, district), 0.0), 4),
            "byTranche": {k: round(v, 4) for k, v in extrap.get((state, district), {}).items() if v > 0},
            "facilityCount": counts[(state, district)],
            "facilities": facs,
        }

    out = {
        "wards": wards,
        "seasonality": seasonality,
        "monthSeason": MONTH_SEASON,
        "monthLabels": MONTH_LABELS,
        "scalars": scalars,
        "tranches": tranches,
        "districts": districts,
    }
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(out, f, indent=1, ensure_ascii=False)

    # Sanity print
    tot = {}
    for st, ds in districts.items():
        tot[st] = round(sum(d["sampledMT"] + sum(d["byTranche"].values()) for d in ds.values()), 1)
    print(f"Wrote {OUT}")
    print(f"  wards={len(wards)} tranches={len(tranches)} states={list(districts.keys())}")
    print(f"  state annual MT (should match Dashboard): {tot}")


if __name__ == "__main__":
    main()
