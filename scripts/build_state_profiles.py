"""
Build-time pipeline for the STATE / DISTRICT costing tab.

Derives, per OXYGEN-BED BAND, a "typical facility archetype" — what oxygen
infrastructure a facility of that size likely has and how hard it runs — from
the WJCF 92-facility survey. State/district users then only enter how many
facilities fall in each band; the engine expands each into its archetype,
applies state unit rates and computes the annual cost.

The raw survey did NOT record facility type (90/92 blank), but it has oxygen-bed
counts, which are a clean size proxy — so bed band is the archetype key.

Output: src/data/state-profiles.json  (aggregate medians only — no facility
names/districts; safe to ship). Rates block mirrors the workbook Assumptions /
Form B defaults; norm-based fields (oximeters, clinical staff, boosters) are not
in the survey and carry documented default norms the user can override.

Run:  python scripts/build_state_profiles.py
"""
import io, json, os, statistics as st, warnings
warnings.filterwarnings("ignore")
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
XLSX = os.path.join(
    ROOT,
    "Facility-Level_Data_Collection__Oxygen_Cost_Tool_-_all_versions_-_English_-_2026-05-27-09-18-28.xlsx",
)
OUT = os.path.join(ROOT, "src", "data", "state-profiles.json")

BANDS = ["<10", "10-29", "30-59", "60+"]
# Human label bridging bed band <-> typical facility level (guidance only).
BAND_LABEL = {
    "<10": "PHC-scale (<10 oxygen beds)",
    "10-29": "CHC-scale (10–29 oxygen beds)",
    "30-59": "SDH-scale (30–59 oxygen beds)",
    "60+": "DH / Medical College-scale (60+ oxygen beds)",
}
# IEC tier per band (Assumptions J1/J2/J3).
BAND_IEC = {"<10": "small", "10-29": "small", "30-59": "mid", "60+": "large"}


def num(v):
    try:
        if v is None or v == "":
            return None
        return float(v)
    except Exception:
        return None


def avg3(r, a, b, c):
    xs = [num(r[a]), num(r[b]), num(r[c])]
    xs = [x for x in xs if x is not None and x > 0]
    return sum(xs) / len(xs) if xs else 0.0


def band_of(b):
    if not b or b <= 0:
        return None
    return "<10" if b < 10 else "10-29" if b < 30 else "30-59" if b < 60 else "60+"


def med(xs, dp=1, default=0.0):
    xs = [x for x in xs if x is not None]
    return round(st.median(xs), dp) if xs else default


def med_pos(xs, dp=1, default=0.0, lo=None, hi=None):
    """Median over positive, in-range values; drops KoBo sentinels 888/999."""
    ys = []
    for x in xs:
        if x is None or x <= 0 or x in (888, 999):
            continue
        if lo is not None and x < lo:
            continue
        if hi is not None and x > hi:
            continue
        ys.append(x)
    return round(st.median(ys), dp) if ys else default


SNAP_LPM = [500, 1000, 2000]


def snap_capacity(cap):
    return min(SNAP_LPM, key=lambda s: abs(s - cap))


# Survey run-hours are unreliable (mixed per-day / monthly / cumulative meter
# readings), so PSA production hours/day uses defensible per-band operating
# defaults instead of the noisy median. Editable in the UI.
PROD_HRS = {"<10": 6, "10-29": 8, "30-59": 10, "60+": 14}
# Dedicated-technician fallback when the survey value is missing/implausible.
TECH_DEFAULT = {"<10": 0, "10-29": 1, "30-59": 2, "60+": 2}


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True, read_only=True)
    M = wb["Facility-Level Data Collecti..."]
    P = wb["psa_repeat"]
    L = wb["lmo_repeat"]

    fac = {}
    for r in M.iter_rows(min_row=2, values_only=True):
        fid = r[81]
        if fid is None:
            continue
        oxbeds = num(r[9]) or num(r[8])
        band = band_of(oxbeds)
        if band is None:
            continue
        fac[fid] = {
            "band": band,
            "state": (str(r[0]).strip() if r[0] else "—"),
            "oxBeds": oxbeds,
            "totalBeds": num(r[6]) or num(r[7]) or oxbeds,
            "funcBeds": num(r[7]) or oxbeds,
            "psaN": num(r[10]) or 0,
            "lmoN": num(r[12]) or 0,
            "cylD": num(r[13]) or 0, "cylB": num(r[14]) or 0, "cylA": num(r[15]) or 0,
            "refD": avg3(r, 16, 17, 18), "refB": avg3(r, 19, 20, 21), "refA": avg3(r, 22, 23, 24),
            "priceD": num(r[25]), "priceB": num(r[26]), "priceA": num(r[27]),
            "distKm": None,
            "ocDep": num(r[41]) or 0, "ocHigh": num(r[42]) or 0, "ocLow": num(r[43]) or 0,
            "mgps": 1 if str(r[50]).strip().lower().startswith("yes") else 0,
            "bhu": num(r[54]) or 0,
            "techN": num(r[60]) or 0, "sal": num(r[61]),
            "_runhrs": [], "_caps": [], "_powerPerLpm": [],
            "_lmoVol": 0.0, "_lmoRefills": 0,
        }

    # PSA plants: production hours/day, capacity, power/LPM (functional only).
    for r in P.iter_rows(min_row=2, values_only=True):
        d = fac.get(r[45])
        if not d:
            continue
        status = (str(r[3]).strip().lower() if r[3] else "")
        if status.startswith("non-functional"):
            continue
        rh = avg3(r, 5, 6, 7)
        cap = num(r[0])
        pw = num(r[1])
        if 0 < rh <= 24:
            d["_runhrs"].append(rh)
        if cap and 100 <= cap <= 3000:
            d["_caps"].append(cap)
        if cap and pw and cap > 0 and 0.02 <= pw / cap <= 0.5:
            d["_powerPerLpm"].append(pw / cap)

    # LMO tanks: annual volume (KL) + refill count.
    for r in L.iter_rows(min_row=2, values_only=True):
        d = fac.get(r[39])
        if not d:
            continue
        vol = avg3(r, 6, 7, 8)  # per-refill or monthly volume (KL-ish); coarse
        if vol > 0:
            d["_lmoVol"] += vol

    prof = {}
    for band in BANDS:
        fs = [d for d in fac.values() if d["band"] == band]
        n = len(fs)
        if n == 0:
            continue
        allcaps = [c for d in fs for c in d["_caps"]]

        def frac(pred):
            return round(sum(1 for d in fs if pred(d)) / n, 2)

        cap = snap_capacity(med(allcaps, 0, 500) or 500)
        techs = med_pos([d["techN"] for d in fs], 0, TECH_DEFAULT[band], lo=1, hi=20)
        lmo_tanks = round(med_pos([d["lmoN"] for d in fs], 0, 1)) or 1
        lmo_kl = 10 if band == "60+" else 5

        prof[band] = {
            "band": band,
            "label": BAND_LABEL[band],
            "n": n,
            "oxBeds": med([d["oxBeds"] for d in fs]),
            "totalBeds": med([d["totalBeds"] for d in fs]),
            "funcBeds": med([d["funcBeds"] for d in fs]),
            "iecTier": BAND_IEC[band],
            # Each *Prob is the share of band facilities that actually have the
            # source — the engine multiplies that source's cost by it, so a band
            # total is the EXPECTED cost across its facilities (budget-accurate).
            # PSA
            "psaProb": frac(lambda d: d["psaN"] > 0),
            "psaPlants": round(med_pos([d["psaN"] for d in fs], 0, 1)) or 1,
            "psaCapacityLpm": cap,
            "psaProdHrsPerDay": PROD_HRS[band],
            # LMO
            "lmoProb": frac(lambda d: d["lmoN"] > 0),
            "lmoTanks": lmo_tanks,
            "lmoCapacityKl": lmo_kl,
            # annual LMO volume (KL): survey volume unreliable; norm = tanks ×
            # capacity × ~monthly refill. Editable preset.
            "lmoAnnualKl": lmo_tanks * lmo_kl * 12,
            # Cylinders (counts available + refills/month; price uses state rate)
            "cylProb": frac(lambda d: (d["cylD"] + d["cylB"] + d["cylA"]) > 0 or (d["refD"] + d["refB"] + d["refA"]) > 0),
            "cylDCount": round(med_pos([d["cylD"] for d in fs], 0, 0, hi=2000)),
            "cylBCount": round(med_pos([d["cylB"] for d in fs], 0, 0, hi=2000)),
            "cylACount": round(med_pos([d["cylA"] for d in fs], 0, 0, hi=1000)),
            "cylDRefillsMo": round(med_pos([d["refD"] for d in fs], 0, 0, hi=300)),
            "cylBRefillsMo": round(med_pos([d["refB"] for d in fs], 0, 0, hi=300)),
            "cylARefillsMo": round(med_pos([d["refA"] for d in fs], 0, 0, hi=300)),
            # Concentrators (deployed units × avg run hours/day; split too noisy)
            "ocProb": frac(lambda d: d["ocDep"] > 0),
            "ocDeployed": round(med_pos([d["ocDep"] for d in fs], 0, 0, hi=500)),
            "ocHrsPerDay": 6,
            # MGPS
            "mgpsProb": frac(lambda d: d["mgps"] > 0),
            "mgpsBhu": round(med_pos([d["bhu"] for d in fs], 0, 0, hi=2000)),
            # HR: total dedicated technicians (split govt/contract unknown ->
            # treated as contractual by default; editable).
            "techProb": frac(lambda d: d["techN"] and 1 <= d["techN"] <= 20),
            "techs": round(techs) or 1,
        }

    # --- Norm-based archetype fields not observed by the survey (editable) ----
    # Oximeters: rule-of-thumb per oxygen bed. Clinical staff for training:
    # per-facility norms scaling with size. Boosters: rare -> 0 by default.
    NORMS = {
        "<10": {"fingertip": 3, "bedside": 2, "doctors": 4, "nurses": 8, "paramedics": 6, "boosters": 0},
        "10-29": {"fingertip": 6, "bedside": 5, "doctors": 8, "nurses": 20, "paramedics": 12, "boosters": 0},
        "30-59": {"fingertip": 12, "bedside": 12, "doctors": 20, "nurses": 45, "paramedics": 25, "boosters": 0},
        "60+": {"fingertip": 25, "bedside": 30, "doctors": 45, "nurses": 110, "paramedics": 60, "boosters": 0},
    }
    for band, p in prof.items():
        p.update(NORMS[band])

    # --- State unit rates (Form B / Assumptions defaults) --------------------
    rates = {
        "electricityTariff": 7.5,               # A1  ₹/kWh
        "psaPowerByCapacity": {"500": 8, "1000": 14, "2000": 26},  # A2-4 kWh/hr
        "ocPowerKwh": 0.3,                       # A5  kWh/hr per concentrator
        "cylRefillD": 120, "cylRefillB": 260, "cylRefillA": 350,   # B1-3 ₹/refill
        "cylTransportPerTrip": 800, "cylPerTrip": 20,              # B4-5
        "cylHydrotest": 300,                      # B6 ₹/cyl (every 5 yrs)
        "lmoRatePerKg": 48,                       # C1 ₹/kg (KL×1000×rate convention)
        "lmoAmcPct": 0.04,                        # C2
        "lmoAssetByKl": {"5": 1200000, "10": 1900000, "20": 2600000},  # C3-4 + F B5.3
        "psaCamcPct": 0.06,                       # D1
        "psaAssetByCapacity": {"500": 2500000, "1000": 4200000, "2000": 7500000},  # D2-4
        "psaRepairPct": 0.035,                    # D5-7 (age unknown -> mid, 5–10y)
        "mgpsAmcPct": 0.04,                       # E1
        "mgpsAssetPerBhu": 15000,                 # E2 ₹/BHU
        "mgpsRepairPct": 0.02,                    # E3
        "ocAmcPct": 0.10,                         # F1
        "ocAsset": 40000,                         # F2 ₹/unit
        "ocFilterPerYear": 800,                   # F3
        "oxiFingertipPerYear": 150,               # G1
        "oxiBedsideProbePerYear": 1200,           # G2
        "oxiBedsideAmcPct": 0.08,                 # G3
        "oxiBedsideAsset": 18000,                 # G4
        "salaryGovtTech": 35000,                  # H1 ₹/mo
        "salaryContractTech": 22000,              # H2 ₹/mo
        "trainDoctor": 1800, "trainNurse": 1200, "trainParamedic": 900,  # I1-3
        "trainPsaTech": 3500,                     # I4
        "refresherEveryYears": 2, "refresherPct": 0.5,  # I5-6
        "iec": {"large": 15000, "mid": 8000, "small": 3000},  # J1-3
        "contingencyPct": 0.10,                   # K1
        # OC daily run hours by intensity (Form A intensity bands) — for OC power
        "ocHighHrs": 12, "ocLowHrs": 4,
    }

    # --- Per-facility infrastructure vectors (anonymized: state + beds + infra,
    # no names/districts) for the runtime k-nearest-neighbours model -----------
    vectors = []
    for d in fac.values():
        caps = [c for c in d["_caps"]]
        cap = snap_capacity(st.median(caps)) if caps else 500
        techN = d["techN"] if (d["techN"] and 1 <= d["techN"] <= 20) else 0
        vectors.append({
            "state": d["state"],
            "oxBeds": round(d["oxBeds"], 1),
            "funcBeds": round(d["funcBeds"] or d["oxBeds"], 1),
            "psa": 1 if d["psaN"] > 0 else 0,
            "psaPlants": int(d["psaN"]) if d["psaN"] > 0 else 0,
            "psaCapacityLpm": cap if d["psaN"] > 0 else 0,
            "lmo": 1 if d["lmoN"] > 0 else 0,
            "lmoTanks": int(d["lmoN"]) if d["lmoN"] > 0 else 0,
            "cyl": 1 if (d["cylD"] + d["cylB"] + d["cylA"] + d["refD"] + d["refB"] + d["refA"]) > 0 else 0,
            "cylDRefillsMo": round(d["refD"]) if 0 < d["refD"] <= 300 else 0,
            "cylBRefillsMo": round(d["refB"]) if 0 < d["refB"] <= 300 else 0,
            "cylARefillsMo": round(d["refA"]) if 0 < d["refA"] <= 300 else 0,
            "cylCount": round(min(d["cylD"] + d["cylB"] + d["cylA"], 4000)),
            "oc": 1 if d["ocDep"] > 0 else 0,
            "ocDeployed": round(d["ocDep"]) if 0 < d["ocDep"] <= 500 else 0,
            "mgps": d["mgps"],
            "bhu": round(d["bhu"]) if 0 < d["bhu"] <= 2000 else 0,
            "techs": techN,
        })

    # --- Per-state rate profiles (only the rates the SURVEY actually observed:
    # cylinder refill prices and per-technician salary). Others stay national. --
    def price_list(state, key):
        return [d[key] for d in fac.values() if (state is None or d["state"] == state)
                and d[key] not in (None, 0, 888, 999) and 20 <= (d[key] or 0) <= 1500]

    def salary_list(state):
        out = []
        for d in fac.values():
            if state is not None and d["state"] != state:
                continue
            if d["sal"] and d["techN"] and 1 <= d["techN"] <= 20:
                per = d["sal"] / d["techN"]
                if 5000 <= per <= 100000:
                    out.append(per)
        return out

    states = sorted({d["state"] for d in fac.values() if d["state"] and d["state"] != "—"})
    state_rates = {}
    for state in states + [None]:
        cd = price_list(state, "priceD")
        cb = price_list(state, "priceB")
        sl = salary_list(state)
        key = state or "All states"
        state_rates[key] = {
            "n": sum(1 for d in fac.values() if state is None or d["state"] == state),
            "cylRefillD": round(st.median(cd)) if cd else None,
            "cylRefillB": round(st.median(cb)) if cb else None,
            "salaryContractTech": round(st.median(sl)) if sl else None,
        }

    state_meta = {
        state: {
            "n": sum(1 for d in fac.values() if d["state"] == state),
            "bedMin": round(min(d["oxBeds"] for d in fac.values() if d["state"] == state)),
            "bedMax": round(max(d["oxBeds"] for d in fac.values() if d["state"] == state)),
        }
        for state in states
    }
    all_beds = [d["oxBeds"] for d in fac.values()]

    out = {
        "meta": {
            "cohortLabel": "WJCF 92-facility oxygen assessment (Madhya Pradesh, Chhattisgarh, Punjab)",
            "n": len(fac),
            "states": state_meta,
            "bedMin": round(min(all_beds)),
            "bedMax": round(max(all_beds)),
            "note": "The model predicts each facility's likely oxygen infrastructure from its oxygen-bed count and state using k-nearest-neighbours over the survey. Oximeters, clinical-staff (training) and booster counts were not surveyed and use documented default norms. All values are user-editable.",
        },
        "bands": [prof[b] for b in BANDS if b in prof],
        "facilities": vectors,
        "stateRates": state_rates,
        "rates": rates,
    }
    io.open(OUT, "w", encoding="utf-8").write(json.dumps(out, indent=1, ensure_ascii=False))
    print("wrote", OUT, "| facilities:", len(vectors), "| states:", states)


if __name__ == "__main__":
    main()
