"""
Build-time data pipeline: convert the WJCF facility-level survey workbook into a
compact, ANONYMIZED JSON knowledge base for the in-app benchmarking features.

Run manually whenever the survey is refreshed:
    python scripts/build_facilities.py

Output: src/data/facilities.json  (anonymized — no facility names or districts).

Anonymization: each facility is reduced to "facility type · bed band · state".
Names (DH names) and districts are dropped.
"""
import io, json, os, statistics as st, warnings
warnings.filterwarnings("ignore")
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
XLSX = os.path.join(
    ROOT,
    "Facility-Level_Data_Collection__Oxygen_Cost_Tool_-_all_versions_-_English_-_2026-05-27-09-18-28.xlsx",
)
OUT = os.path.join(ROOT, "src", "data", "facilities.json")

COHORT = "WJCF facility-level oxygen assessment — 92 facilities across Madhya Pradesh, Chhattisgarh & Punjab"
PERIOD = "Nov 2025 – Jan 2026"

# conversion factors (mirror the engine)
LMO_EXPANSION = 0.861
KG_TO_CUM = 0.7
CYL = {"d": 7.0, "b": 1.5, "a": 0.66}
COMP_RUN = 0.9  # compressor-run fraction used to derive PSA output from run hrs


def num(v):
    try:
        if v is None or v == "":
            return None
        return float(v)
    except Exception:
        return None


def avg3(r, a, b, c):
    xs = [num(r[a]), num(r[b]), num(r[c])]
    xs = [x for x in xs if x is not None and x > 0]
    return sum(xs) / len(xs) if xs else 0.0


def bed_band(b):
    if not b or b <= 0:
        return None
    if b < 10:
        return "<10"
    if b < 30:
        return "10–29"
    if b < 60:
        return "30–59"
    return "60+"


def clean(xs, lo, hi, drop=()):
    return [x for x in xs if x is not None and lo <= x <= hi and x not in drop]


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True, read_only=True)
    M = wb["Facility-Level Data Collecti..."]
    P = wb["psa_repeat"]
    L = wb["lmo_repeat"]

    mrows = list(M.iter_rows(min_row=2, values_only=True))

    fac = {}
    for r in mrows:
        fid = r[81]  # _id
        if fid is None:
            continue
        ftype = (str(r[72]).strip() if r[72] else "") or "Facility"
        oxbeds = num(r[9]) or num(r[8]) or None
        cylD = avg3(r, 16, 17, 18) or (num(r[74]) or 0)
        cylB = avg3(r, 19, 20, 21) or (num(r[75]) or 0)
        cylA = avg3(r, 22, 23, 24) or (num(r[76]) or 0)
        cyl_spend = avg3(r, 30, 31, 32)  # monthly, all types
        fac[fid] = {
            "state": str(r[0]).strip() if r[0] else "—",
            "facilityType": ftype,
            "oxBeds": oxbeds,
            "bedBand": bed_band(oxbeds),
            "psaCount": int(num(r[10]) or 0),
            "lmoCount": int(num(r[12]) or 0),
            "cylD": cylD, "cylB": cylB, "cylA": cylA,
            "ocDeployed": int(num(r[41]) or 0),
            "ocHigh": int(num(r[42]) or 0),
            "ocLow": int(num(r[43]) or 0),
            "cylRefillD": num(r[25]),
            "cylRefillB": num(r[26]),
            "hrSalary": num(r[61]),
            "cylSpend": cyl_spend,
            # accumulators
            "_psaOut": 0.0, "_lmoOut": 0.0,
            "_lmoSpend": 0.0, "_lmoRentalVals": [],
            "_psaPowerPerLpm": [],
        }

    # PSA plants
    psa_total = 0
    psa_nonfunc = 0
    for r in P.iter_rows(min_row=2, values_only=True):
        psa_total += 1
        status = str(r[3]).strip() if r[3] else ""
        if status.lower().startswith("non-functional"):
            psa_nonfunc += 1
        fid = r[45]  # _submission__id
        d = fac.get(fid)
        if not d:
            continue
        cap = num(r[0]) or 0
        pw = num(r[1])
        run = avg3(r, 5, 6, 7)
        d["_psaOut"] += run * COMP_RUN * cap * 60 / 1000
        if cap and pw and cap > 0:
            d["_psaPowerPerLpm"].append(pw / cap)

    # LMO tanks  (join via _submission__id = col 39)
    UNIT = {
        "Nm³ (Normal cubic metres)": 1.0,
        "KL (Kilolitres)": 1000 * LMO_EXPANSION,
        "Kg (Kilograms)": KG_TO_CUM,
    }
    for r in L.iter_rows(min_row=2, values_only=True):
        fid = r[39]
        d = fac.get(fid)
        if not d:
            continue
        vol = avg3(r, 6, 7, 8)
        unit = str(r[9]) if r[9] else ""
        factor = UNIT.get(unit, 1.0)  # default: treat as cu m
        d["_lmoOut"] += vol * factor
        d["_lmoSpend"] += avg3(r, 16, 17, 18) + avg3(r, 19, 20, 21)
        rent = num(r[23])
        if rent and rent > 1000:
            d["_lmoRentalVals"].append(rent)

    facilities = []
    rentals_all = []
    perCuM = {"psa": [], "lmo": [], "cylinder": []}
    for i, (fid, d) in enumerate(fac.items()):
        cyl_out = d["cylD"] * CYL["d"] + d["cylB"] * CYL["b"] + d["cylA"] * CYL["a"]
        oc_out = (d["ocHigh"] * 10 + d["ocLow"] * 4) * 30 * 5 * 60 / 1000
        outs = {
            "psa": round(d["_psaOut"], 1),
            "lmo": round(d["_lmoOut"], 1),
            "cylinder": round(cyl_out, 1),
            "oc": round(oc_out, 1),
        }
        total = sum(outs.values())
        # Primary = the dominant BULK / medical-grade source (PSA, LMO or
        # cylinders). Concentrators are supplementary (low-purity) and never the
        # "primary" supply, so they are excluded from this choice and tracked
        # separately via sources.oc.
        bulk = {k: outs[k] for k in ("psa", "lmo", "cylinder")}
        primary = max(bulk, key=bulk.get) if sum(bulk.values()) > 0 else None

        # actual per-cu-m where computable
        cyl_cpm = round(d["cylSpend"] / cyl_out, 2) if cyl_out > 0 and d["cylSpend"] > 0 else None
        lmo_rent = st.median(d["_lmoRentalVals"]) if d["_lmoRentalVals"] else 0
        lmo_cpm = (
            round((d["_lmoSpend"] + lmo_rent) / d["_lmoOut"], 2)
            if d["_lmoOut"] > 0 and (d["_lmoSpend"] + lmo_rent) > 0
            else None
        )
        if cyl_cpm and 1 <= cyl_cpm <= 500:
            perCuM["cylinder"].append(cyl_cpm)
        if lmo_cpm and 1 <= lmo_cpm <= 500:
            perCuM["lmo"].append(lmo_cpm)
        for v in d["_lmoRentalVals"]:
            rentals_all.append(v)

        psa_ppl = st.median(d["_psaPowerPerLpm"]) if d["_psaPowerPerLpm"] else None

        facilities.append({
            "id": f"F{i + 1:02d}",
            "state": d["state"],
            "facilityType": d["facilityType"],
            "oxBeds": d["oxBeds"],
            "bedBand": d["bedBand"],
            "sources": {
                "psa": d["psaCount"] > 0,
                "lmo": d["lmoCount"] > 0,
                "cylinder": (d["cylD"] + d["cylB"] + d["cylA"]) > 0,
                "oc": d["ocDeployed"] > 0,
            },
            "output": {**outs, "total": round(total, 1)},
            "primary": primary,
            "perCuM": {"cylinder": cyl_cpm, "lmo": lmo_cpm, "psa": None},
            "metrics": {
                "psaPowerPerLpm": round(psa_ppl, 4) if psa_ppl else None,
                "cylRefillD": d["cylRefillD"] if d["cylRefillD"] not in (None, 888) else None,
                "cylRefillB": d["cylRefillB"] if d["cylRefillB"] not in (None, 888) else None,
                "hrSalary": d["hrSalary"],
            },
        })

    # distributions for outlier flags / percentiles / ranges (sorted, cleaned)
    dist = {
        "psaPowerPerLpm": sorted(clean([f["metrics"]["psaPowerPerLpm"] for f in facilities], 0.02, 0.5)),
        "cylRefillD": sorted(clean([f["metrics"]["cylRefillD"] for f in facilities], 50, 1500, drop=(888,))),
        "cylRefillB": sorted(clean([f["metrics"]["cylRefillB"] for f in facilities], 20, 1500, drop=(888,))),
        "hrSalary": sorted(clean([f["metrics"]["hrSalary"] for f in facilities], 1000, 500000)),
        "lmoRental": sorted(clean(rentals_all, 1000, 300000)),
        "perCuM": {k: sorted(v) for k, v in perCuM.items()},
    }

    states = {}
    for f in facilities:
        states[f["state"]] = states.get(f["state"], 0) + 1

    out = {
        "meta": {
            "cohortLabel": COHORT,
            "period": PERIOD,
            "facilityCount": len(facilities),
            "states": states,
            "psaPlants": {
                "total": psa_total,
                "nonFunctional": psa_nonfunc,
                "nonFunctionalPct": round(psa_nonfunc / psa_total * 100) if psa_total else 0,
            },
        },
        "facilities": facilities,
        "distributions": dist,
    }

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    io.open(OUT, "w", encoding="utf-8", newline="\n").write(json.dumps(out, ensure_ascii=False, indent=1))
    print("Wrote", OUT)
    print("facilities:", len(facilities), "| states:", states)
    print("primary dist:", {p: sum(1 for f in facilities if f["primary"] == p) for p in ["psa", "lmo", "cylinder", "oc", None]})
    print("PSA non-functional:", psa_nonfunc, "/", psa_total)
    print("cost coverage n:", {k: len(v) for k, v in dist["perCuM"].items()})
    print("flag dist n:", {k: len(v) for k, v in dist.items() if k != "perCuM"})


if __name__ == "__main__":
    main()
